"""
app/agents/catalog_bridge_agent.py
────────────────────────────────────────────────────────────────────────────────
Nyx Bridge — Puente de catálogo e inventario Shein → ZOPA dinámica para Hermes.

Fuentes (prioridad):
  1. Snapshot JSON en cache local
  2. Excel shein_enterizos_deportivos.xlsx
  3. Defaults de margen configurables por env
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL = _PROJECT_ROOT / "shein_enterizos_deportivos.xlsx"
CACHE_JSON = Path(os.environ.get("LOCAL_STORAGE_DIR", "app/storage_vault")) / "catalog_snapshot.json"

DEFAULT_MARGEN_PCT = float(os.environ.get("CATALOG_MARGEN_PCT", "120"))
DEFAULT_COSTO_ENVIO = float(os.environ.get("CATALOG_COSTO_ENVIO", "8000"))
DEFAULT_MARGEN_MINIMO = float(os.environ.get("CATALOG_MARGEN_MINIMO", "15000"))


@dataclass
class CatalogProduct:
    goods_id: str
    titulo: str
    precio_cop: float
    precio_cop_texto: str
    imagen_url: str
    producto_url: str
    precio_reventa: float = 0.0
    target_price: float = 0.0
    reserve_price: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        return d


class CatalogBridgeAgent:
    """Lee snapshots Shein y calcula límites ZOPA por producto."""

    def __init__(
        self,
        excel_path: Path | str | None = None,
        margen_pct: float = DEFAULT_MARGEN_PCT,
        costo_envio: float = DEFAULT_COSTO_ENVIO,
    ):
        self.excel_path = Path(excel_path) if excel_path else DEFAULT_EXCEL
        self.margen_pct = margen_pct
        self.costo_envio = costo_envio
        self._products: list[CatalogProduct] = []
        self._loaded_at: Optional[str] = None
        self.refresh()

    def refresh(self) -> int:
        """Recarga catálogo desde JSON cache o Excel."""
        if CACHE_JSON.exists():
            try:
                data = json.loads(CACHE_JSON.read_text(encoding="utf-8"))
                self._products = [self._dict_to_product(p) for p in data.get("products", [])]
                self._loaded_at = data.get("updated_at")
                if self._products:
                    return len(self._products)
            except Exception as e:
                print(f"[CatalogBridge] cache JSON error: {e}")

        if self.excel_path.exists():
            try:
                self._products = self._load_from_excel(self.excel_path)
            except PermissionError:
                print(
                    f"[CatalogBridge] Excel bloqueado ({self.excel_path.name}) — "
                    "cierra el archivo o usa POST /agents/catalog/refresh"
                )
                self._products = [self._default_product()]
            except Exception as e:
                print(f"[CatalogBridge] Error leyendo Excel: {e}")
                self._products = [self._default_product()]
            self._loaded_at = datetime.now(timezone.utc).isoformat()
            if len(self._products) > 1 or self._products[0].goods_id != "default":
                self._persist_cache()
            return len(self._products)

        self._products = [self._default_product()]
        self._loaded_at = datetime.now(timezone.utc).isoformat()
        return len(self._products)

    def get_products(self, limit: int = 50) -> list[dict[str, Any]]:
        return [p.to_dict() for p in self._products[:limit]]

    def get_featured_product(self) -> Optional[dict[str, Any]]:
        if not self._products:
            return None
        # Producto con mejor equilibrio precio/ventas: el de precio mediano
        sorted_p = sorted(self._products, key=lambda x: x.precio_cop)
        mid = sorted_p[len(sorted_p) // 2]
        return mid.to_dict()

    def get_top_seller(self, query: str = "") -> dict[str, Any]:
        """
        Producto prioritario para campañas / content pipeline.
        1. Match por query (product_focus, niche, keyword)
        2. Primer producto con imagen y mejor precio_reventa (margen)
        3. Featured / default
        """
        if query.strip():
            found = self.find_product(query)
            if found:
                return found

        candidates = [
            p for p in self._products
            if p.goods_id != "default" and (p.imagen_url or p.producto_url)
        ]
        if candidates:
            best = max(candidates, key=lambda p: p.precio_reventa)
            return best.to_dict()

        return self.get_featured_product() or self._default_product().to_dict()

    def find_product(self, query: str) -> Optional[dict[str, Any]]:
        if not query or not self._products:
            return None
        q = _normalize(query)
        tokens = [t for t in q.split() if len(t) > 2]

        best: Optional[CatalogProduct] = None
        best_score = 0
        for p in self._products:
            title = _normalize(p.titulo)
            score = sum(1 for t in tokens if t in title)
            if score > best_score:
                best_score = score
                best = p
        if best and best_score > 0:
            return best.to_dict()

        # Match numérico de precio en mensaje
        price_match = re.search(r"(\d{2,3}[.\s]?\d{3})", query.replace(",", ""))
        if price_match:
            target = _parse_cop(price_match.group(1))
            closest = min(self._products, key=lambda p: abs(p.precio_reventa - target))
            return closest.to_dict()
        return None

    def get_zopa_for_message(self, user_message: str) -> dict[str, float]:
        """Resuelve producto por mensaje o usa featured/default."""
        product = self.find_product(user_message)
        if not product:
            product = self.get_featured_product() or self._default_product().to_dict()
        return {
            "target_price": product["target_price"],
            "reserve_price": product["reserve_price"],
            "precio_reventa": product["precio_reventa"],
            "precio_shein": product["precio_cop"],
            "titulo": product["titulo"],
        }

    def get_default_zopa(self) -> dict[str, float]:
        p = self.get_featured_product() or self._default_product().to_dict()
        return {
            "target_price": p["target_price"],
            "reserve_price": p["reserve_price"],
        }

    def get_catalog_summary(self) -> dict[str, Any]:
        precios = [p.precio_reventa for p in self._products if p.precio_reventa > 0]
        return {
            "total_products": len(self._products),
            "loaded_at": self._loaded_at,
            "avg_reventa": round(sum(precios) / len(precios), 0) if precios else 0,
            "price_range": {
                "min": min(precios) if precios else 0,
                "max": max(precios) if precios else 0,
            },
            "top_titles": [p.titulo for p in self._products[:5]],
        }

    def ingest_shein_products(self, products: list[Any]) -> int:
        """Importa lista SheinProduct del scraper."""
        self._products = []
        for item in products:
            if hasattr(item, "to_dict"):
                raw = item.to_dict()
            elif isinstance(item, dict):
                raw = item
            else:
                continue
            self._products.append(self._raw_to_product(raw))
        self._loaded_at = datetime.now(timezone.utc).isoformat()
        self._persist_cache()
        return len(self._products)

    def _load_from_excel(self, path: Path) -> list[CatalogProduct]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[CatalogBridge] openpyxl no instalado")
            return [self._default_product()]

        products: list[CatalogProduct] = []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            titulo = row[1]
            precio_raw = row[2]
            if not titulo:
                continue
            precio = _parse_cop(precio_raw)
            if precio <= 0:
                continue
            raw = {
                "goods_id": str(row[0] or ""),
                "titulo": str(titulo),
                "precio_cop": precio,
                "precio_cop_texto": str(precio_raw or ""),
                "imagen_url": str(row[3] or ""),
                "producto_url": str(row[4] or ""),
            }
            products.append(self._raw_to_product(raw))
        wb.close()
        return products or [self._default_product()]

    def _raw_to_product(self, raw: dict[str, Any]) -> CatalogProduct:
        precio = float(raw.get("precio_cop") or 0)
        zopa = _calc_zopa(precio, self.margen_pct, self.costo_envio)
        return CatalogProduct(
            goods_id=str(raw.get("goods_id") or ""),
            titulo=str(raw.get("titulo") or "Producto"),
            precio_cop=precio,
            precio_cop_texto=str(raw.get("precio_cop_texto") or ""),
            imagen_url=str(raw.get("imagen_url") or ""),
            producto_url=str(raw.get("producto_url") or ""),
            precio_reventa=zopa["precio_reventa"],
            target_price=zopa["target_price"],
            reserve_price=zopa["reserve_price"],
        )

    def _dict_to_product(self, d: dict[str, Any]) -> CatalogProduct:
        return CatalogProduct(
            goods_id=str(d.get("goods_id", "")),
            titulo=str(d.get("titulo", "")),
            precio_cop=float(d.get("precio_cop", 0)),
            precio_cop_texto=str(d.get("precio_cop_texto", "")),
            imagen_url=str(d.get("imagen_url", "")),
            producto_url=str(d.get("producto_url", "")),
            precio_reventa=float(d.get("precio_reventa", 0)),
            target_price=float(d.get("target_price", 0)),
            reserve_price=float(d.get("reserve_price", 0)),
        )

    def _default_product(self) -> CatalogProduct:
        zopa = _calc_zopa(45990, self.margen_pct, self.costo_envio)
        return CatalogProduct(
            goods_id="default",
            titulo="Enterizo deportivo trending",
            precio_cop=45990,
            precio_cop_texto="COP$ 45.990",
            imagen_url="",
            producto_url="",
            precio_reventa=zopa["precio_reventa"],
            target_price=zopa["target_price"],
            reserve_price=zopa["reserve_price"],
        )

    def _persist_cache(self) -> None:
        CACHE_JSON.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "updated_at": self._loaded_at,
            "products": [p.to_dict() for p in self._products],
            "summary": self.get_catalog_summary(),
        }
        CACHE_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


_bridge_instance: Optional[CatalogBridgeAgent] = None


def get_catalog_bridge() -> CatalogBridgeAgent:
    global _bridge_instance
    if _bridge_instance is None:
        _bridge_instance = CatalogBridgeAgent()
    return _bridge_instance


def _calc_zopa(
    precio_shein: float, margen_pct: float, costo_envio: float
) -> dict[str, float]:
    precio_reventa = round(precio_shein * (1 + margen_pct / 100), 0)
    reserve = round(precio_shein + costo_envio + DEFAULT_MARGEN_MINIMO, 0)
    target = precio_reventa
    if reserve >= target:
        target = round(reserve * 1.15, 0)
    return {
        "precio_reventa": precio_reventa,
        "target_price": target,
        "reserve_price": reserve,
    }


def _parse_cop(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).upper().replace("COP", "").replace("$", "").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        m = re.search(r"(\d+)", s.replace(".", ""))
        return float(m.group(1)) if m else 0.0


def _normalize(text: str) -> str:
    t = text.lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        t = t.replace(a, b)
    return t
