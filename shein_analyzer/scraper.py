"""
shein_analyzer/scraper.py

Scraper Playwright para Shein Colombia — categoría Enterizos Deportivos.
Extrae título, precio en COP, enlace de imagen y enlace del producto.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import random
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote, urljoin

from dotenv import load_dotenv
from playwright.async_api import BrowserContext, Page, Response, async_playwright

load_dotenv()

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Configuración Shein Colombia
# ─────────────────────────────────────────────
BASE_URL = "https://co.shein.com"
CATEGORY_KEYWORD = "enterizos deportivos"
# sort=8 → Más popular / más vendidos
SORT_MOST_SOLD = "8"
DEFAULT_LIMIT = 30
DEFAULT_TIMEOUT_MS = 60_000
DEFAULT_USER_DATA_DIR = ".shein_browser_profile"
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILENAME = "shein_enterizos_deportivos.xlsx"
EXCEL_OUTPUT_PATH = PROJECT_ROOT / EXCEL_FILENAME

# User-Agent de Chrome actualizado en Windows (debe coincidir con sec-ch-ua)
REAL_CHROME_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)
CHROME_MAJOR_VERSION = "132"
DEFAULT_CAPTCHA_WAIT_S = 60
COP_CURRENCY_FORMAT = '#,##0 "COP"'

_STEALTH_INIT_SCRIPT = """
(() => {
  // Ocultar webdriver y señales de automatización
  Object.defineProperty(navigator, 'webdriver', { get: () => false });
  try {
    delete Object.getPrototypeOf(navigator).webdriver;
  } catch (e) {}

  if (!window.chrome) {
    window.chrome = { runtime: {}, loadTimes: () => ({}), csi: () => ({}) };
  }

  Object.defineProperty(navigator, 'language', {
    get: () => 'es-CO',
  });

  Object.defineProperty(navigator, 'languages', {
    get: () => ['es-CO', 'es'],
  });

  Object.defineProperty(navigator, 'plugins', {
    get: () => [1, 2, 3, 4, 5],
  });

  Object.defineProperty(navigator, 'platform', {
    get: () => 'Win32',
  });

  const originalQuery = window.navigator.permissions.query;
  window.navigator.permissions.query = (parameters) =>
    parameters.name === 'notifications'
      ? Promise.resolve({ state: Notification.permission })
      : originalQuery(parameters);
})();
"""

_BOT_CHALLENGE_HINTS = (
    "soy humano",
    "i'm human",
    "im human",
    "verify you are human",
    "verification required",
    "captcha",
    "geetest",
    "slider",
    "robot",
)

_API_URL_HINTS = (
    "product-list",
    "productList",
    "pdsearch",
    "getSearchInfo",
    "CategoryProduct",
    "goods_list",
    "bff-api",
)

_COOKIE_ACCEPT_SELECTORS = (
    "#onetrust-accept-btn-handler",
    "button:has-text('Accept All')",
    "button:has-text('Allow all')",
    "button:has-text('Aceptar todo')",
    "button:has-text('Confirm my choices')",
    "button:has-text('Confirmar mis elecciones')",
)

_PRODUCT_CARD_SELECTORS = (
    "[class*='product-card']",
    "[data-testid='product-card']",
    ".S-product-list__item",
    "[class*='product-list__item']",
    "a[href*='-p-']",
)

_PRICE_RENDER_SELECTORS = (
    "[class*='product-item__c-price']",
    "[class*='product-card__price']",
    "[class*='goods-price']",
    "[class*='sale-price']",
    "[class*='c-price']",
    "[class*='price-box']",
)


@dataclass
class SheinProduct:
    """Producto normalizado extraído de Shein Colombia."""

    goods_id: str
    titulo: str
    precio_cop: float
    precio_cop_texto: str
    imagen_url: str
    producto_url: str
    ventas: int = 0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class SheinScraper:
    """
    Scraper Playwright para listados de Shein Colombia.

    Flujo:
      1. Navega a la búsqueda de enterizos deportivos (sort=8).
      2. Espera explícitamente a que aparezcan tarjetas de producto.
      3. Intercepta respuestas JSON de la API interna de Shein.
      4. Usa el DOM como respaldo si la API no devuelve datos a tiempo.
    """

    def __init__(
        self,
        *,
        headless: bool = False,
        timeout_ms: int = DEFAULT_TIMEOUT_MS,
        locale: str = "es-CO",
        user_data_dir: Optional[str] = None,
        use_chrome_channel: bool = True,
    ):
        self.headless = headless
        self.timeout_ms = timeout_ms
        self.locale = locale
        self.user_data_dir = user_data_dir or os.getenv(
            "SHEIN_USER_DATA_DIR", DEFAULT_USER_DATA_DIR
        )
        self.use_chrome_channel = use_chrome_channel
        self._captured_payloads: list[Any] = []

    @property
    def category_url(self) -> str:
        keyword = quote(CATEGORY_KEYWORD)
        return (
            f"{BASE_URL}/pdsearch/{keyword}/"
            f"?sort={SORT_MOST_SOLD}&page=1&limit=120"
        )

    async def scrape_products(self, limit: int = DEFAULT_LIMIT) -> list[SheinProduct]:
        """Extrae los productos principales de enterizos deportivos."""
        self._captured_payloads.clear()

        async with async_playwright() as playwright:
            context = await self._create_stealth_context(playwright)
            page = context.pages[0] if context.pages else await context.new_page()
            page.set_default_timeout(self.timeout_ms)
            page.on("response", self._on_response)

            try:
                await self._warm_up_session(page)
                await self._navigate_like_human(page, self.category_url)

                if await self._is_bot_challenge(page):
                    await self._wait_out_bot_challenge(page)

                await self._dismiss_cookie_banner(page)
                await self._human_warmup_before_scrape(page)
                await self._wait_for_product_cards(page)
                await self._scroll_to_load_more(page, passes=5)
                await self._wait_for_prices_rendered(page)

                products = self._parse_api_payloads(limit)
                if len(products) < limit:
                    dom_products = await self._parse_dom_cards(page, limit)
                    products = self._merge_products(products, dom_products, limit)

                if not products and await self._is_bot_challenge(page):
                    logger.error(
                        "Shein bloqueó el scraping con verificación 'Soy humano'. "
                        "Ejecuta con SHEIN_HEADLESS=false y resuelve el CAPTCHA manualmente."
                    )

                products.sort(key=lambda item: item.ventas, reverse=True)
                return products[:limit]
            finally:
                await context.close()

    async def _create_stealth_context(self, playwright) -> BrowserContext:
        """Crea un contexto persistente con flags anti-detección."""
        profile_dir = Path(self.user_data_dir).resolve()
        profile_dir.mkdir(parents=True, exist_ok=True)

        launch_kwargs: dict[str, Any] = {
            "headless": self.headless,
            "locale": self.locale,
            "timezone_id": "America/Bogota",
            "viewport": {"width": 1366, "height": 768},
            "user_agent": REAL_CHROME_UA,
            "color_scheme": "light",
            "device_scale_factor": 1,
            "has_touch": False,
            "is_mobile": False,
            "java_script_enabled": True,
            "accept_downloads": False,
            "ignore_https_errors": False,
            "extra_http_headers": {
                "Accept-Language": "es-CO,es;q=0.9,en-US,en;q=0.8",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,image/apng,*/*;q=0.8"
                ),
                "sec-ch-ua": (
                    f'"Google Chrome";v="{CHROME_MAJOR_VERSION}", '
                    f'"Chromium";v="{CHROME_MAJOR_VERSION}", '
                    '"Not_A Brand";v="24"'
                ),
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
                "Upgrade-Insecure-Requests": "1",
                "DNT": "1",
            },
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process,AutomationControlled",
                "--disable-infobars",
                "--disable-dev-shm-usage",
                "--no-first-run",
                "--no-default-browser-check",
                "--window-size=1366,768",
            ],
            "ignore_default_args": ["--enable-automation"],
            "slow_mo": random.randint(40, 120),
        }

        if self.use_chrome_channel:
            launch_kwargs["channel"] = "chrome"

        try:
            context = await playwright.chromium.launch_persistent_context(
                str(profile_dir),
                **launch_kwargs,
            )
        except Exception as exc:
            if "channel" in launch_kwargs:
                logger.warning(
                    "Chrome instalado no disponible (%s). Usando Chromium de Playwright.",
                    exc,
                )
                launch_kwargs.pop("channel", None)
                context = await playwright.chromium.launch_persistent_context(
                    str(profile_dir),
                    **launch_kwargs,
                )
            else:
                raise

        await context.add_init_script(_STEALTH_INIT_SCRIPT)
        return context

    async def _warm_up_session(self, page: Page) -> None:
        """Visita la home primero para obtener cookies y parecer un usuario real."""
        logger.info("Calentando sesión en homepage: %s", BASE_URL)
        await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=self.timeout_ms)
        await self._random_pause(2.0, 4.0)
        await self._dismiss_cookie_banner(page)
        await self._simulate_human_behavior(page, light=True)

    async def _navigate_like_human(self, page: Page, url: str) -> None:
        """Navega a la categoría con pausas y movimiento de mouse."""
        logger.info("Navegando a categoría: %s", url)
        await self._random_pause(0.8, 1.6)
        await page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
        await self._random_pause(1.5, 3.0)
        await self._simulate_human_behavior(page)

    async def _human_warmup_before_scrape(self, page: Page) -> None:
        """Pausas humanas y scroll progresivo antes de extraer datos."""
        logger.info("Simulando navegación humana antes de extraer productos...")
        await self._random_pause(2.0, 4.0)
        await self._simulate_human_behavior(page)
        await self._progressive_scroll(page, passes=5)
        await self._random_pause(2.0, 4.0)

    async def _progressive_scroll(self, page: Page, passes: int = 4) -> None:
        """Scroll gradual con mouse wheel imitando lectura de catálogo."""
        for step in range(passes):
            delta = random.randint(180, 420)
            await page.mouse.wheel(0, delta)
            await self._random_pause(0.6, 1.4)

            await page.mouse.move(
                random.randint(120, 1100),
                random.randint(160, 640),
                steps=random.randint(12, 28),
            )

            if step > 0 and step % 2 == 0:
                await page.mouse.wheel(0, -random.randint(60, 140))
                await self._random_pause(0.3, 0.7)

    async def _simulate_human_behavior(self, page: Page, *, light: bool = False) -> None:
        """Movimientos suaves de cursor y micro-scroll."""
        viewport = page.viewport_size or {"width": 1366, "height": 768}
        width = viewport["width"]
        height = viewport["height"]

        moves = 2 if light else random.randint(4, 8)
        for _ in range(moves):
            x = random.randint(80, max(120, width - 120))
            y = random.randint(80, max(120, height - 120))
            await page.mouse.move(x, y, steps=random.randint(14, 32))
            await self._random_pause(0.2, 0.55)

        if not light:
            await page.mouse.wheel(0, random.randint(140, 360))
            await self._random_pause(0.5, 1.2)

    async def _is_bot_challenge(self, page: Page) -> bool:
        """Detecta CAPTCHA, pantalla 'Soy humano' o URL risk/challenge."""
        current_url = page.url.lower()
        if "risk/challenge" in current_url or "/challenge" in current_url:
            return True

        try:
            body_text = (await page.locator("body").inner_text(timeout=5_000)).lower()
        except Exception:
            body_text = (await page.content()).lower()

        return any(hint in body_text for hint in _BOT_CHALLENGE_HINTS)

    async def _wait_out_bot_challenge(self, page: Page) -> None:
        """
        Espera hasta 60 s a que el usuario resuelva el CAPTCHA manualmente.
        """
        max_wait_s = int(os.getenv("SHEIN_CAPTCHA_WAIT_S", str(DEFAULT_CAPTCHA_WAIT_S)))
        elapsed = 0
        interval = 3

        while elapsed < max_wait_s:
            if not await self._is_bot_challenge(page):
                logger.info("Desafío anti-bot superado tras %s s", elapsed)
                await self._random_pause(1.0, 2.0)
                return

            if elapsed == 0:
                _alert_captcha_detected(page.url)

            await asyncio.sleep(interval)
            elapsed += interval

            await page.mouse.move(
                random.randint(200, 600),
                random.randint(200, 500),
                steps=random.randint(10, 22),
            )

        logger.warning("Timeout esperando CAPTCHA (%s s)", max_wait_s)

    @staticmethod
    async def _random_pause(min_s: float, max_s: float) -> None:
        await asyncio.sleep(random.uniform(min_s, max_s))

    async def _on_response(self, response: Response) -> None:
        url = response.url
        if response.status != 200:
            return
        if not any(hint in url for hint in _API_URL_HINTS):
            return

        content_type = (response.headers.get("content-type") or "").lower()
        if "json" not in content_type and not url.endswith(".json"):
            return

        try:
            payload = await response.json()
            self._captured_payloads.append(payload)
        except Exception:
            return

    async def _dismiss_cookie_banner(self, page: Page) -> None:
        """Cierra el banner de cookies si aparece."""
        for selector in _COOKIE_ACCEPT_SELECTORS:
            try:
                button = page.locator(selector).first
                if await button.is_visible(timeout=2_000):
                    await button.click(timeout=3_000)
                    await page.wait_for_timeout(800)
                    return
            except Exception:
                continue

    async def _wait_for_product_cards(self, page: Page) -> None:
        """Espera explícita a que carguen tarjetas de producto."""
        for selector in _PRODUCT_CARD_SELECTORS:
            try:
                await page.wait_for_selector(selector, timeout=12_000)
                logger.info("Productos detectados con selector: %s", selector)
                return
            except Exception:
                continue

        # Espera implícita adicional por contenido dinámico
        await page.wait_for_timeout(3_000)

    async def _wait_for_prices_rendered(self, page: Page) -> None:
        """Espera a que los precios COP se pinten en las tarjetas."""
        for selector in _PRICE_RENDER_SELECTORS:
            try:
                await page.wait_for_selector(selector, timeout=10_000)
                logger.info("Precios detectados con selector: %s", selector)
                await self._random_pause(0.8, 1.5)
                return
            except Exception:
                continue

        try:
            await page.get_by_text(re.compile(r"COP\$", re.I)).first.wait_for(timeout=8_000)
            logger.info("Precios detectados por texto COP$")
            await self._random_pause(0.8, 1.5)
            return
        except Exception:
            pass

        logger.warning("No se detectaron precios renderizados; se intentará parseo parcial.")
        await page.wait_for_timeout(2_000)

    async def _scroll_to_load_more(self, page: Page, passes: int = 3) -> None:
        """Scroll humano adicional para lazy-load de productos y precios."""
        await self._progressive_scroll(page, passes=passes)

    def _parse_api_payloads(self, limit: int) -> list[SheinProduct]:
        raw_items: list[dict[str, Any]] = []
        seen_ids: set[str] = set()

        for payload in self._captured_payloads:
            for item in _walk_product_dicts(payload):
                goods_id = _extract_goods_id(item)
                if not goods_id or goods_id in seen_ids:
                    continue
                seen_ids.add(goods_id)
                raw_items.append(item)
                if len(raw_items) >= limit * 2:
                    break

        products = [_normalize_product(item) for item in raw_items]
        return [product for product in products if product is not None]

    async def _parse_dom_cards(self, page: Page, limit: int) -> list[SheinProduct]:
        """Respaldo: extrae datos visibles del DOM con selectores de Shein."""
        cards = await page.evaluate(
            """(maxItems) => {
                const priceRegex = /(?:COP\\$|COL\\$|\\$)\\s*([\\d][\\d.,]*)/i;
                const salesRegex = /(\\d[\\d.,]*)\\+?\\s*(?:vendidos|sold|ventas|soldes)/i;

                const titleSelectors = [
                    "[class*='goods-title']",
                    "[class*='product-item__goods-title']",
                    "[class*='product-card__goods-title']",
                    "[class*='product-item__title']",
                    "[class*='title-content']",
                    "[class*='goods-name']",
                    "[class*='product-name']",
                    ".S-product-card__name",
                ];

                const priceSelectors = [
                    "[class*='product-item__c-price']",
                    "[class*='product-card__price']",
                    "[class*='goods-price']",
                    "[class*='sale-price']",
                    "[class*='discount-price']",
                    "[class*='price__sale']",
                    "[class*='c-price']:not([class*='original']):not([class*='orginal'])",
                    "[class*='price-box'] [class*='price']",
                ];

                function isLikelyTitle(text) {
                    if (!text) return false;
                    const t = text.trim();
                    if (t.length < 8) return false;
                    if (/^\\d+$/.test(t)) return false;
                    if (/^\\d+(\\.\\d+)?$/.test(t)) return false;
                    return /[a-zA-ZáéíóúñÁÉÍÓÚÑ]/.test(t);
                }

                function titleFromUrl(href) {
                    const match = href.match(/\\/([^/?#]+)-p-\\d+\\.html/i);
                    if (!match) return "";
                    return match[1]
                        .replace(/-/g, " ")
                        .replace(/\\s+/g, " ")
                        .trim();
                }

                function extractTitle(card, anchor, href) {
                    for (const sel of titleSelectors) {
                        const el = card.querySelector(sel);
                        if (!el) continue;
                        const candidate = (
                            el.getAttribute("title")
                            || el.getAttribute("aria-label")
                            || el.textContent
                            || ""
                        ).replace(/\\s+/g, " ").trim();
                        if (isLikelyTitle(candidate)) return candidate.slice(0, 220);
                    }

                    const img = card.querySelector("img[alt]");
                    if (img && isLikelyTitle(img.alt)) return img.alt.trim().slice(0, 220);

                    const aria = anchor?.getAttribute("aria-label") || "";
                    if (isLikelyTitle(aria)) return aria.trim().slice(0, 220);

                    const anchorTitle = anchor?.getAttribute("title") || "";
                    if (isLikelyTitle(anchorTitle)) return anchorTitle.trim().slice(0, 220);

                    return titleFromUrl(href);
                }

                function extractPrice(card) {
                    for (const sel of priceSelectors) {
                        const nodes = card.querySelectorAll(sel);
                        for (const node of nodes) {
                            const text = (node.textContent || "").replace(/\\s+/g, " ").trim();
                            const match = text.match(priceRegex);
                            if (match) {
                                return {
                                    precio_texto: text.match(/(?:COP\\$|COL\\$|\\$)\\s*[\\d.,]+/i)?.[0] || text,
                                    precio_num: match[1],
                                };
                            }
                        }
                    }

                    const priceNodes = card.querySelectorAll("span, div, p, strong, b");
                    for (const node of priceNodes) {
                        const text = (node.textContent || "").replace(/\\s+/g, " ").trim();
                        if (!text || text.length > 24) continue;
                        const match = text.match(/^(?:COP\\$|COL\\$|\\$)\\s*[\\d.,]+$/i);
                        if (match) {
                            const numMatch = text.match(/([\\d.,]+)/);
                            return {
                                precio_texto: text,
                                precio_num: numMatch ? numMatch[1] : "",
                            };
                        }
                    }

                    const cardText = (card.innerText || "").replace(/\\s+/g, " ");
                    const textMatch = cardText.match(priceRegex);
                    if (textMatch) {
                        return {
                            precio_texto: textMatch[0],
                            precio_num: textMatch[1],
                        };
                    }

                    return { precio_texto: "", precio_num: "" };
                }

                function findCard(anchor) {
                    return (
                        anchor.closest(".S-product-list__item")
                        || anchor.closest("[class*='product-list__item']")
                        || anchor.closest("[class*='product-card']")
                        || anchor.closest("[class*='product-item']")
                        || anchor.closest("[class*='goods-card']")
                        || anchor.closest("li")
                        || anchor.closest("article")
                        || anchor.parentElement
                    );
                }

                const anchors = Array.from(document.querySelectorAll("a[href*='-p-']"));
                const seen = new Set();
                const results = [];

                for (const anchor of anchors) {
                    const href = anchor.href || "";
                    const idMatch = href.match(/-p-(\\d+)\\.html/i);
                    if (!idMatch || seen.has(idMatch[1])) continue;

                    const card = findCard(anchor);
                    if (!card) continue;

                    const titulo = extractTitle(card, anchor, href);
                    if (!titulo) continue;

                    seen.add(idMatch[1]);

                    const price = extractPrice(card);
                    const cardText = (card.innerText || "").replace(/\\s+/g, " ");
                    const salesMatch = cardText.match(salesRegex);

                    const img = card.querySelector("img");
                    const imgSrc =
                        img?.src
                        || img?.dataset?.src
                        || img?.getAttribute("data-before-crop-src")
                        || img?.getAttribute("data-src")
                        || "";

                    results.push({
                        goods_id: idMatch[1],
                        titulo,
                        precio_texto: price.precio_texto,
                        precio_num: price.precio_num,
                        ventas_texto: salesMatch ? salesMatch[1] : "0",
                        imagen: imgSrc,
                        url: href.split("?")[0] + (href.includes("?") ? "?" + href.split("?")[1].split("&").slice(0, 2).join("&") : ""),
                    });

                    if (results.length >= maxItems) break;
                }

                return results;
            }""",
            limit * 2,
        )

        products: list[SheinProduct] = []
        for card in cards:
            titulo = _sanitize_title(card.get("titulo") or "", card.get("url", ""))
            if not titulo:
                continue

            precio_texto = (card.get("precio_texto") or "").strip()
            precio = _parse_cop_amount(card.get("precio_num") or precio_texto)
            if precio <= 0 and precio_texto:
                precio = _parse_cop_amount(precio_texto)

            products.append(
                SheinProduct(
                    goods_id=str(card.get("goods_id", "")),
                    titulo=titulo,
                    precio_cop=precio,
                    precio_cop_texto=_format_cop_texto(precio, precio_texto),
                    ventas=_parse_sales_count(card.get("ventas_texto", "0")),
                    imagen_url=_normalize_image_url(card.get("imagen", "")),
                    producto_url=card.get("url") or "",
                )
            )

        return products

    @staticmethod
    def _merge_products(
        primary: list[SheinProduct],
        secondary: list[SheinProduct],
        limit: int,
    ) -> list[SheinProduct]:
        merged: dict[str, SheinProduct] = {item.goods_id: item for item in primary}
        for product in secondary:
            if product.goods_id not in merged:
                merged[product.goods_id] = product
                continue

            existing = merged[product.goods_id]
            if _is_better_title(product.titulo, existing.titulo):
                existing.titulo = product.titulo
            if product.precio_cop > 0 and existing.precio_cop <= 0:
                existing.precio_cop = product.precio_cop
                existing.precio_cop_texto = product.precio_cop_texto
            if product.ventas > existing.ventas:
                existing.ventas = product.ventas
            if not existing.imagen_url and product.imagen_url:
                existing.imagen_url = product.imagen_url
            if not existing.producto_url and product.producto_url:
                existing.producto_url = product.producto_url

        return list(merged.values())[: limit * 2]


# ─────────────────────────────────────────────
# Helpers de parsing
# ─────────────────────────────────────────────
def _title_from_url(url: str) -> str:
    """Deriva el nombre del producto desde el slug de la URL de Shein."""
    match = re.search(r"/([^/?#]+)-p-\d+\.html", url, re.IGNORECASE)
    if not match:
        return ""
    return match.group(1).replace("-", " ").strip()


def _is_better_title(candidate: str, current: str) -> bool:
    """Prefiere títulos descriptivos sobre ratings numéricos."""
    candidate = (candidate or "").strip()
    current = (current or "").strip()
    if not candidate:
        return False
    if not current:
        return True
    if re.fullmatch(r"\d+", current) and not re.fullmatch(r"\d+", candidate):
        return True
    return len(candidate) > len(current) and len(candidate) >= 12


def _sanitize_title(raw_title: str, product_url: str) -> str:
    """Limpia títulos erróneos (ratings) y usa fallback desde URL."""
    title = (raw_title or "").strip()
    if re.fullmatch(r"\d+", title) or len(title) < 8:
        title = _title_from_url(product_url)
    return title.strip()


def _format_cop_texto(amount: float, raw_text: str = "") -> str:
    if raw_text and _parse_cop_amount(raw_text) > 0:
        return raw_text.strip()
    if amount <= 0:
        return "COP$ 0"
    formatted = f"{amount:,.0f}".replace(",", ".")
    return f"COP$ {formatted}"


def _walk_product_dicts(node: Any) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []

    def visit(obj: Any) -> None:
        if isinstance(obj, dict):
            if _looks_like_product(obj):
                found.append(obj)
            for value in obj.values():
                visit(value)
        elif isinstance(obj, list):
            for item in obj:
                visit(item)

    visit(node)
    return found


def _looks_like_product(obj: dict[str, Any]) -> bool:
    keys = set(obj.keys())
    id_keys = {"goods_id", "goodsId", "product_id", "productId", "goods_sn"}
    name_keys = {"goods_name", "goodsName", "productName", "title"}
    price_keys = {"salePrice", "retailPrice", "discountPrice", "price"}
    return bool(keys & id_keys) and bool(keys & name_keys) and bool(keys & price_keys)


def _extract_goods_id(item: dict[str, Any]) -> Optional[str]:
    for key in ("goods_id", "goodsId", "product_id", "productId", "goods_sn"):
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


def _normalize_product(item: dict[str, Any]) -> Optional[SheinProduct]:
    goods_id = _extract_goods_id(item)
    if not goods_id:
        return None

    titulo = (
        item.get("goods_name")
        or item.get("goodsName")
        or item.get("productName")
        or item.get("title")
        or ""
    ).strip()
    if not titulo:
        return None

    price_block = (
        item.get("salePrice")
        or item.get("discountPrice")
        or item.get("retailPrice")
        or item.get("price")
        or {}
    )

    if isinstance(price_block, dict):
        precio_texto = str(
            price_block.get("amountWithSymbol")
            or price_block.get("usdAmountWithSymbol")
            or price_block.get("amount")
            or ""
        )
        precio_cop = _parse_cop_amount(precio_texto or price_block.get("amount"))
    else:
        precio_texto = str(price_block)
        precio_cop = _parse_cop_amount(precio_texto)

    imagen = (
        item.get("goods_img")
        or item.get("goodsImg")
        or item.get("detail_image")
        or item.get("goods_thumb")
        or item.get("image")
        or ""
    )
    if isinstance(imagen, list):
        imagen = imagen[0] if imagen else ""

    producto_url = (
        item.get("productUrl")
        or item.get("goods_url")
        or item.get("detail_url")
        or item.get("url")
        or ""
    )
    if producto_url and not str(producto_url).startswith("http"):
        producto_url = urljoin(BASE_URL, str(producto_url))

    ventas = _parse_sales_count(
        item.get("soldNum")
        or item.get("sold_num")
        or item.get("salesVolume")
        or item.get("soldCount")
        or 0
    )

    titulo = _sanitize_title(titulo, str(producto_url))

    return SheinProduct(
        goods_id=goods_id,
        titulo=titulo,
        precio_cop=precio_cop,
        precio_cop_texto=_format_cop_texto(precio_cop, precio_texto),
        ventas=ventas,
        imagen_url=_normalize_image_url(str(imagen)),
        producto_url=str(producto_url),
    )


def _parse_cop_amount(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    cleaned = re.sub(r"[^\d.,]", "", text)
    if not cleaned:
        return 0.0

    # Formato colombiano: 45.990 / 1.234.567 (punto como separador de miles)
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", cleaned):
        return float(cleaned.replace(".", ""))
    if re.fullmatch(r"\d{1,3}(,\d{3})+", cleaned):
        return float(cleaned.replace(",", ""))

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        parts = cleaned.split(",")
        cleaned = cleaned.replace(",", ".") if len(parts[-1]) <= 2 else cleaned.replace(",", "")

    try:
        amount = float(cleaned)
        # Si parece decimal pero en COP suele ser miles (ej. 45.99 → revisar contexto)
        if amount < 1000 and "." in str(value) and re.search(r"\.\d{3}", str(value)):
            return float(re.sub(r"[^\d]", "", str(value)))
        return amount
    except ValueError:
        return 0.0


def _parse_sales_count(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value).lower().strip()
    if not text:
        return 0

    multiplier = 1
    if "k" in text or "mil" in text:
        multiplier = 1_000
        text = text.replace("k", "").replace("mil", "")

    digits = re.sub(r"[^\d]", "", text)
    return int(digits) * multiplier if digits else 0


def _normalize_image_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if url.startswith("//"):
        return f"https:{url}"
    if url.startswith("/"):
        return urljoin(BASE_URL, url)
    return url


# ─────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────
async def scrape_enterizos_deportivos(
    limit: int = DEFAULT_LIMIT,
    *,
    headless: bool = False,
    save_excel: bool = True,
) -> list[dict[str, Any]]:
    """Atajo async que devuelve dicts listos para JSON o pandas."""
    scraper = SheinScraper(headless=headless)
    products = await scraper.scrape_products(limit=limit)
    if save_excel and products:
        excel_path = save_products_to_excel(products)
        logger.info("Excel guardado en: %s", excel_path)
    return [product.to_dict() for product in products]


def _alert_captcha_detected(page_url: str = "") -> None:
    """Emite aviso sonoro y mensaje explícito en consola."""
    banner = "!" * 68
    print(f"\n{banner}")
    print("  CAPTCHA / VERIFICACION DETECTADA — Resuelve 'Soy humano' en el navegador")
    print(f"  URL actual: {page_url or 'desconocida'}")
    print(f"  Esperando hasta {DEFAULT_CAPTCHA_WAIT_S} segundos para resolucion manual...")
    print(f"{banner}\n", flush=True)

    if sys.platform == "win32":
        try:
            import winsound

            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception:
            print("\a", end="", flush=True)
    else:
        print("\a", end="", flush=True)


def _autofit_worksheet_columns(
    worksheet,
    *,
    min_width: float = 8,
    max_width: float = 85,
) -> None:
    """Ajusta el ancho de columnas según el contenido."""
    from openpyxl.utils import get_column_letter

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(
            max(longest + 2, min_width),
            max_width,
        )


def save_products_to_excel(
    products: list[SheinProduct],
    output_path: Optional[Path] = None,
) -> Path:
    """
    Guarda los productos extraídos en un archivo Excel con formato profesional.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if not products:
        raise ValueError("No hay productos para exportar a Excel.")

    destination = Path(output_path or EXCEL_OUTPUT_PATH).resolve()
    extracted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Enterizos Deportivos"

    headers = (
        "#",
        "Título del Producto",
        "Precio (COP)",
        "Enlace de la Imagen (URL)",
        "Enlace del Producto (URL)",
        "Fecha de Extracción",
    )

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_font = Font(bold=True, color="1F2937", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    worksheet.row_dimensions[1].height = 28

    link_font = Font(color="0563C1", underline="single")
    text_alignment = Alignment(vertical="top", wrap_text=True)
    first_data_row = 2

    for position, product in enumerate(products, start=1):
        row = first_data_row + position - 1

        worksheet.cell(row=row, column=1, value=position).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        worksheet.cell(row=row, column=2, value=product.titulo).alignment = text_alignment

        price_cell = worksheet.cell(
            row=row,
            column=3,
            value=float(product.precio_cop) if product.precio_cop > 0 else None,
        )
        price_cell.number_format = COP_CURRENCY_FORMAT
        price_cell.alignment = Alignment(horizontal="right", vertical="top")

        image_cell = worksheet.cell(row=row, column=4, value=product.imagen_url)
        image_cell.alignment = text_alignment
        if product.imagen_url:
            image_cell.hyperlink = product.imagen_url
            image_cell.font = link_font

        product_cell = worksheet.cell(row=row, column=5, value=product.producto_url)
        product_cell.alignment = text_alignment
        if product.producto_url:
            product_cell.hyperlink = product.producto_url
            product_cell.font = link_font

        worksheet.cell(row=row, column=6, value=extracted_at).alignment = Alignment(
            horizontal="center", vertical="top"
        )

    last_product_row = first_data_row + len(products) - 1
    prices = [float(product.precio_cop) for product in products if product.precio_cop > 0]

    if prices:
        summary_row = last_product_row + 2
        avg_price = sum(prices) / len(prices)
        min_price = min(prices)
        max_price = max(prices)

        summary_labels = {
            3: "Precio Promedio",
            4: "Precio Mínimo",
            5: "Precio Máximo",
        }

        title_cell = worksheet.cell(row=summary_row, column=2, value="RESUMEN DE PRECIOS (COP)")
        title_cell.font = summary_font
        title_cell.fill = summary_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        for column_index, value in ((3, avg_price), (4, min_price), (5, max_price)):
            label_cell = worksheet.cell(
                row=summary_row - 1,
                column=column_index,
                value=summary_labels[column_index],
            )
            label_cell.font = Font(bold=True, color="1F2937")
            label_cell.fill = summary_fill
            label_cell.alignment = Alignment(horizontal="center", vertical="center")

            value_cell = worksheet.cell(row=summary_row, column=column_index, value=value)
            value_cell.number_format = COP_CURRENCY_FORMAT
            value_cell.font = summary_font
            value_cell.fill = summary_fill
            value_cell.alignment = Alignment(horizontal="right", vertical="center")

        worksheet.cell(row=summary_row, column=1, value="").fill = summary_fill
        worksheet.cell(row=summary_row, column=6, value=extracted_at).fill = summary_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:F{last_product_row}"
    _autofit_worksheet_columns(worksheet)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _print_results(products: list[SheinProduct]) -> None:
    sep = "-" * 78
    scraper = SheinScraper()

    print(f"\n{sep}")
    print("  SHEIN ANALYZER — Enterizos Deportivos (Colombia)")
    print(f"  URL: {scraper.category_url}")
    print(sep)
    print(f"  {'#':<3} {'PRECIO':<16} {'TITULO':<55}")
    print(f"  {'-'*3} {'-'*15} {'-'*54}")

    for index, product in enumerate(products, 1):
        print(
            f"  {index:<3} {product.precio_cop_texto:<16} "
            f"{product.titulo[:54]:<55}"
        )

    print(sep)
    print("\n  DETALLE (top 5):\n")

    for index, product in enumerate(products[:5], 1):
        print(f"  #{index} {product.titulo}")
        print(f"      Precio : {product.precio_cop_texto}")
        print(f"      Imagen : {product.imagen_url}")
        print(f"      URL    : {product.producto_url}")
        print()

    print(f"{sep}\n  Total extraídos: {len(products)}\n{sep}\n")


def _print_excel_saved(path: Path) -> None:
    print(f"\n  [OK] Archivo Excel guardado en:\n       {path}\n")


async def main() -> list[SheinProduct]:
    limit = int(os.getenv("SHEIN_SCRAPE_LIMIT", str(DEFAULT_LIMIT)))
    # Por defecto visible: Shein suele bloquear headless con CAPTCHA
    headless = os.getenv("SHEIN_HEADLESS", "false").lower() == "true"
    use_chrome = os.getenv("SHEIN_USE_CHROME", "true").lower() != "false"

    scraper = SheinScraper(headless=headless, use_chrome_channel=use_chrome)
    products = await scraper.scrape_products(limit=limit)
    _print_results(products)

    if products:
        excel_path = save_products_to_excel(products)
        _print_excel_saved(excel_path)
    else:
        print("\n  [AVISO] No se generó Excel porque no se extrajeron productos.\n")

    return products


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")

    if sys.platform == "win32":
        import io

        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    asyncio.run(main())
